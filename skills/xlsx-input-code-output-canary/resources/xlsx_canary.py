#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "openpyxl>=3.1.5",
#   "python-calamine>=0.4.0",
# ]
# ///
from __future__ import annotations

import argparse, hashlib, json, os, re, tempfile, urllib.request, zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:  # pragma: no cover - dependency import is surfaced at runtime.
    Workbook = None
    load_workbook = None
    DefinedName = None
    Table = None
    TableStyleInfo = None

try:
    from python_calamine import CalamineWorkbook
except Exception:  # pragma: no cover - dependency import is surfaced in reports.
    CalamineWorkbook = None

CATEGORY_VALUES = {"simple", "complex"}
CLASS_VALUES = {"data", "calculation", "mixed", "guarded", "unsupported"}
ACE_WORKBOOK_CLASS_VALUES = {"data_workbook", "calculation_workbook", "report_workbook", "excluded_workbook"}
REQUIRED_FIELDS = {
    "public_fixture_id",
    "category",
    "workbook",
    "url",
    "license",
    "bytes",
    "sha256",
    "purpose",
    "expected_flags",
}
EXPECTED_FIXTURE_IDS = {
    "S1-base-xlsx",
    "S2-any-sheets-xlsx",
    "S3-with-table-xlsx",
    "S4-merge-cells-xlsx",
    "S5-date-1904-xlsx",
    "C1-cross-sheet-xlsx",
    "C2-defined-names-xlsx",
    "C3-structured-table-refs-xlsx",
    "C4-charts-xlsx",
    "C5-protected-workbook-xlsx",
}
SHA_RE = re.compile(r"^[a-f0-9]{64}$")


def repo_root(start: Path) -> Path | None:
    cur = start.resolve()
    for parent in [cur, *cur.parents]:
        if (parent / ".git").exists():
            return parent
    return None


def load_json(path: Path) -> Any:
    return json.loads(path.read_text())


def write_json(data: Any) -> None:
    print(json.dumps(data, indent=2, sort_keys=True, default=str))


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def validate_manifest(path: Path) -> dict[str, Any]:
    data = load_json(path)
    errors: list[str] = []
    if data.get("canary_manifest_version") != 1:
        errors.append("canary_manifest_version must be 1")
    workbooks = data.get("workbooks")
    if not isinstance(workbooks, list) or len(workbooks) != 10:
        errors.append("manifest must contain exactly 10 workbooks")
        workbooks = workbooks if isinstance(workbooks, list) else []
    seen: set[str] = set()
    counts = {"simple": 0, "complex": 0}
    for i, item in enumerate(workbooks):
        missing = sorted(REQUIRED_FIELDS - set(item))
        if missing:
            errors.append(f"workbooks[{i}] missing fields: {missing}")
            continue
        fixture_id = item["public_fixture_id"]
        if fixture_id in seen:
            errors.append(f"duplicate public fixture id: {fixture_id}")
        seen.add(fixture_id)
        if fixture_id not in EXPECTED_FIXTURE_IDS:
            errors.append(f"workbooks[{i}] public_fixture_id must be a known public canary fixture id")
        if item["category"] not in CATEGORY_VALUES:
            errors.append(f"{fixture_id}: category must be one of {sorted(CATEGORY_VALUES)}")
        else:
            counts[item["category"]] += 1
        if not str(item["url"]).startswith("https://"):
            errors.append(f"{fixture_id}: url must be https")
        if not SHA_RE.match(str(item["sha256"])):
            errors.append(f"{fixture_id}: sha256 must be 64 lowercase hex chars")
        if not isinstance(item["bytes"], int) or item["bytes"] <= 0:
            errors.append(f"{fixture_id}: bytes must be a positive integer")
        if not item["license"]:
            errors.append(f"{fixture_id}: license must be recorded")
        if not isinstance(item["expected_flags"], list) or not item["expected_flags"]:
            errors.append(f"{fixture_id}: expected_flags must be a non-empty list")
    if counts != {"simple": 5, "complex": 5}:
        errors.append(f"manifest must have 5 simple and 5 complex entries, got {counts}")
    if seen != EXPECTED_FIXTURE_IDS:
        errors.append("manifest public_fixture_id set must match the known public canary fixture ids")
    return {"ok": not errors, "errors": errors, "counts": counts, "public_fixture_ids": sorted(seen)}


def default_cache_dir() -> Path:
    base = Path(os.environ.get("XDG_CACHE_HOME", Path.home() / ".cache"))
    return base / "raw-to-knowledge-playbook" / "xlsx-canary"


def ensure_cache_allowed(cache_dir: Path, allow_repo_cache: bool) -> None:
    root = repo_root(Path.cwd())
    resolved = cache_dir.resolve()
    if root and (resolved == root or root in resolved.parents) and not allow_repo_cache:
        raise SystemExit("refusing to write workbook bytes under repo; use an external cache dir")


def fetch_manifest(manifest: Path, cache_dir: Path, allow_repo_cache: bool) -> dict[str, Any]:
    status = validate_manifest(manifest)
    if not status["ok"]:
        return status
    ensure_cache_allowed(cache_dir, allow_repo_cache)
    cache_dir.mkdir(parents=True, exist_ok=True)
    results = []
    for item in load_json(manifest)["workbooks"]:
        target = cache_dir / f"{item['public_fixture_id']}-{item['workbook']}"
        with urllib.request.urlopen(item["url"], timeout=30) as response:
            payload = response.read()
        target.write_bytes(payload)
        actual_sha = sha256_file(target)
        result = {
            "public_fixture_id": item["public_fixture_id"],
            "path": str(target),
            "bytes": len(payload),
            "sha256": actual_sha,
            "ok": len(payload) == item["bytes"] and actual_sha == item["sha256"],
        }
        if result["ok"]:
            try:
                evidence = inventory_workbook(target, item["public_fixture_id"])
                flags = observed_flags(evidence)
                missing = sorted(set(item["expected_flags"]) - flags)
                result.update({
                    "classification": classify_inventory(evidence)["classification"],
                    "observed_flags": sorted(flags),
                    "missing_expected_flags": missing,
                    "ok": not missing,
                })
            except Exception as exc:
                result.update({"ok": False, "inventory_error": str(exc)})
        results.append(result)
    return {"ok": all(r["ok"] for r in results), "cache_dir": str(cache_dir), "results": results}


def defined_name_titles(wb: Any) -> list[str]:
    names = wb.defined_names
    values = names.values() if hasattr(names, "values") else names.definedName
    return sorted(name.name for name in values)


def workbook_is_protected(wb: Any) -> bool:
    security = getattr(wb, "security", None)
    if security is None:
        return False
    return bool(getattr(security, "lockStructure", False) or getattr(security, "lockWindows", False))


def inventory_sheet(ws: Any, value_ws: Any) -> dict[str, Any]:
    formulas = []
    non_empty = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                non_empty += 1
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append({
                    "cell": cell.coordinate,
                    "formula": cell.value,
                    "cached_value": value_ws[cell.coordinate].value,
                })
    return {
        "name": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "formula_count": len(formulas),
        "formulas": formulas[:50],
        "tables": sorted(ws.tables.keys()),
        "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
        "chart_count": len(getattr(ws, "_charts", [])),
        "sheet_protected": bool(ws.protection.sheet),
        "non_empty_cells": non_empty,
    }


def inventory_workbook(path: Path, public_fixture_id: str | None = None) -> dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to inventory workbook bytes")
    suffix = path.suffix.lower()
    keep_vba = suffix == ".xlsm"
    formula_wb = load_workbook(path, data_only=False, read_only=False, keep_links=True, keep_vba=keep_vba)
    value_wb = load_workbook(path, data_only=True, read_only=True, keep_links=True, keep_vba=keep_vba)
    sheets = []
    for ws in formula_wb.worksheets:
        sheets.append(inventory_sheet(ws, value_wb[ws.title]))
    calamine_names: list[str] = []
    if CalamineWorkbook is not None:
        calamine_names = CalamineWorkbook.from_path(path).sheet_names
    return {
        "public_fixture_id": public_fixture_id,
        "workbook": path.name,
        "sha256": sha256_file(path),
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "parser_versions": parser_versions(),
        "sheet_names": [ws.title for ws in formula_wb.worksheets],
        "calamine_sheet_names": calamine_names,
        "date_system": str(formula_wb.epoch.date()),
        "defined_names": defined_name_titles(formula_wb),
        "workbook_protected": workbook_is_protected(formula_wb),
        "file_extension": suffix,
        "macro_presence": workbook_has_macro_payload(path),
        "external_link_count": len(getattr(formula_wb, "_external_links", []) or []),
        "sheets": sheets,
        "summary": {
            "formula_count": sum(s["formula_count"] for s in sheets),
            "table_count": sum(len(s["tables"]) for s in sheets),
            "merged_range_count": sum(len(s["merged_ranges"]) for s in sheets),
            "chart_count": sum(s["chart_count"] for s in sheets),
            "protected_sheet_count": sum(1 for s in sheets if s["sheet_protected"]),
        },
    }


def parser_versions() -> dict[str, str]:
    import openpyxl
    try:
        import python_calamine
        calamine_version = getattr(python_calamine, "__version__", "unknown")
    except Exception:
        calamine_version = "unavailable"
    return {"openpyxl": openpyxl.__version__, "python_calamine": calamine_version}


def classify_inventory(evidence: dict[str, Any]) -> dict[str, Any]:
    summary = evidence["summary"]
    reasons = []
    guarded = evidence["workbook_protected"] or summary["protected_sheet_count"] > 0
    if guarded:
        classification = "guarded"
        status = "deferred"
        reasons.append("workbook or sheet protection is enabled")
    elif summary["formula_count"] > 0:
        classification = "mixed" if summary["table_count"] or summary["chart_count"] else "calculation"
        status = "requires_calculation_proof"
        reasons.append("formula cells are present; cached values are evidence only")
        if summary["chart_count"]:
            reasons.append("chart objects are present and require non-tabular inventory")
        if summary["table_count"]:
            reasons.append("native tables are present and require structured output handling")
    elif summary["chart_count"] > 0:
        classification = "unsupported"
        status = "deferred"
        reasons.append("charts are present without a semantic proof lane")
    else:
        classification = "data"
        status = "ready_for_parser_output"
        reasons.append("no formulas, charts, or protection detected")
    triplet = triplet_contract(classification)
    return {
        "classification": classification,
        "status": status,
        "reasons": reasons,
        "triplet": triplet,
        "valid_classification": classification in CLASS_VALUES,
    }


def ace_classification_from_inventory(evidence: dict[str, Any]) -> dict[str, Any]:
    original = classify_inventory(evidence)
    original_class = original["classification"]
    summary = evidence.get("summary", {})
    extension = str(evidence.get("file_extension") or Path(str(evidence.get("workbook", ""))).suffix).lower()
    flags = ace_inventory_flags(evidence)
    deferrals: list[str] = []
    report_evidence = _has_report_evidence(summary)

    if extension in {".xls", ".xlsb", ".ods"}:
        workbook_class = "excluded_workbook"
        route_target = "metadata_only"
        deferrals.append("adapter_required")
    elif original_class == "guarded":
        workbook_class = "excluded_workbook"
        route_target = "excluded_no_ingest"
        deferrals.append("protected_workbook")
    elif original_class in {"calculation", "mixed"}:
        workbook_class = "calculation_workbook"
        route_target = "metadata_only"
    elif original_class == "unsupported" and report_evidence:
        workbook_class = "report_workbook"
        route_target = "metadata_only"
    elif original_class == "data":
        workbook_class = "data_workbook"
        route_target = "metadata_only"
    else:
        workbook_class = "excluded_workbook"
        route_target = "excluded_no_ingest"
        deferrals.append("unsupported_workbook")

    if "macro_present" in flags or "external_links_present" in flags:
        deferrals.append("macro_or_external_logic_deferred")

    return {
        "original_canary_class": original_class,
        "workbook_class": workbook_class,
        "route_target": route_target,
        "content_eligible": workbook_class != "excluded_workbook",
        "status": original["status"],
        "inventory_flags": flags,
        "deferral_reasons": sorted(set(deferrals)),
        "report_evidence": report_evidence,
        "valid_workbook_class": workbook_class in ACE_WORKBOOK_CLASS_VALUES,
    }


def ace_inventory_flags(evidence: dict[str, Any]) -> list[str]:
    flags = sorted(observed_flags(evidence))
    if evidence.get("macro_presence"):
        flags.append("macro_present")
    if int(evidence.get("external_link_count") or 0) > 0:
        flags.append("external_links_present")
    return sorted(set(flags))


def workbook_has_macro_payload(path: Path) -> bool:
    if path.suffix.lower() != ".xlsm" or not zipfile.is_zipfile(path):
        return False
    with zipfile.ZipFile(path) as archive:
        return any(name.lower() == "xl/vbaproject.bin" for name in archive.namelist())


def _has_report_evidence(summary: dict[str, Any]) -> bool:
    chart_count = int(summary.get("chart_count") or 0)
    return chart_count > 0


def observed_flags(evidence: dict[str, Any]) -> set[str]:
    summary = evidence["summary"]
    sheets = evidence.get("sheets", [])
    sheet_names = evidence.get("sheet_names") or [
        sheet["name"] for sheet in sheets if isinstance(sheet, dict) and "name" in sheet
    ]
    defined_names = evidence.get("defined_names") or []
    flags: set[str] = set()
    if summary["formula_count"]:
        flags.add("formula")
    if len(sheet_names) > 1:
        flags.add("multiple_sheets")
    if summary["table_count"]:
        flags.add("table")
    if summary["merged_range_count"]:
        flags.add("merged_cells")
    if str(evidence.get("date_system", "")).startswith("1904"):
        flags.add("date_1904")
    if defined_names:
        flags.add("defined_names")
    if summary["chart_count"]:
        flags.add("charts")
    if evidence.get("workbook_protected") or summary["protected_sheet_count"]:
        flags.add("protection")
    if not (summary["formula_count"] or summary["chart_count"] or "protection" in flags):
        flags.add("data_only")
    for sheet in sheets:
        if any("!" in item["formula"] for item in sheet.get("formulas", [])):
            flags.add("cross_sheet_formula")
    return flags


def triplet_contract(classification: str) -> dict[str, str]:
    contracts = {
        "data": ("sheet/range schema", "parser or schema code", "normalized table output"),
        "calculation": ("input cells plus formula graph", "evaluator adapter or ported formula code", "recomputed output proof"),
        "mixed": ("separated data ranges plus formula graph", "parser plus evaluator/code path", "normalized data and recomputed output proof"),
        "guarded": ("guard/protection evidence", "no code until access/review is approved", "explicit deferral artifact"),
        "unsupported": ("unsupported feature evidence", "no code until lane exists", "explicit deferral artifact"),
    }
    input_contract, code_artifact, output_artifact = contracts[classification]
    return {
        "input_contract": input_contract,
        "code_artifact": code_artifact,
        "output_artifact": output_artifact,
    }


def make_simple_workbook(path: Path) -> None:
    if Workbook is None or Table is None or TableStyleInfo is None:
        raise RuntimeError("openpyxl is required for workbook self-test")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    rows = [["date", "value"], [datetime(2026, 1, 1), 10], [datetime(2026, 1, 2), 20]]
    for row in rows:
        ws.append(row)
    table = Table(displayName="DataTable", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    wb.save(path)


def make_complex_workbook(path: Path) -> None:
    if Workbook is None or DefinedName is None:
        raise RuntimeError("openpyxl is required for workbook self-test")
    wb = Workbook()
    inputs = wb.active
    inputs.title = "Inputs"
    inputs["A1"] = 100
    inputs["A2"] = 0.05
    calc = wb.create_sheet("Calc")
    calc["A1"] = "=Inputs!A1*(1+Inputs!A2)"
    calc["A2"] = "=Principal+10"
    wb.defined_names.add(DefinedName("Principal", attr_text="'Inputs'!$A$1"))
    wb.save(path)


def make_guarded_workbook(path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for workbook self-test")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "guarded"
    wb.security.lockStructure = True
    ws.protection.sheet = True
    wb.save(path)


def run_self_test() -> dict[str, Any]:
    with tempfile.TemporaryDirectory() as td:
        root = Path(td)
        simple = root / "simple.xlsx"
        complex_ = root / "complex.xlsx"
        guarded = root / "guarded.xlsx"
        make_simple_workbook(simple)
        make_complex_workbook(complex_)
        make_guarded_workbook(guarded)
        simple_evidence = inventory_workbook(simple, "generated-simple")
        complex_evidence = inventory_workbook(complex_, "generated-complex")
        guarded_evidence = inventory_workbook(guarded, "generated-guarded")
        checks = [
            classify_inventory(simple_evidence)["classification"] == "data",
            simple_evidence["summary"]["table_count"] == 1,
            classify_inventory(complex_evidence)["status"] == "requires_calculation_proof",
            complex_evidence["summary"]["formula_count"] == 2,
            bool(complex_evidence["defined_names"]),
            classify_inventory(guarded_evidence)["classification"] == "guarded",
        ]
        return {"ok": all(checks), "checks": checks, "parser_versions": parser_versions()}


def main() -> int:
    parser = argparse.ArgumentParser(description="XLS/XLSX canary manifest, inventory, and classification helper.")
    sub = parser.add_subparsers(dest="command", required=True)
    m = sub.add_parser("manifest-check")
    m.add_argument("--manifest", type=Path, required=True)
    f = sub.add_parser("fetch")
    f.add_argument("--manifest", type=Path, required=True)
    f.add_argument("--cache-dir", type=Path, default=default_cache_dir())
    f.add_argument("--allow-repo-cache", action="store_true")
    i = sub.add_parser("inventory")
    i.add_argument("--workbook", type=Path, required=True)
    i.add_argument("--public-fixture-id")
    i.add_argument("--source-id", dest="public_fixture_id")
    c = sub.add_parser("classify")
    c.add_argument("--inventory", type=Path, required=True)
    c.add_argument("--ace", action="store_true", help="emit ACE wave-2 workbook class mapping")
    sub.add_parser("self-test")
    args = parser.parse_args()
    if args.command == "manifest-check":
        result = validate_manifest(args.manifest)
    elif args.command == "fetch":
        result = fetch_manifest(args.manifest, args.cache_dir, args.allow_repo_cache)
    elif args.command == "inventory":
        result = inventory_workbook(args.workbook, args.public_fixture_id)
    elif args.command == "classify":
        inventory = load_json(args.inventory)
        result = ace_classification_from_inventory(inventory) if args.ace else classify_inventory(inventory)
    else:
        result = run_self_test()
    write_json(result)
    return 0 if result.get("ok", True) else 1


if __name__ == "__main__":
    raise SystemExit(main())
